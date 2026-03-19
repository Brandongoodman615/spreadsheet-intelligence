"""
Generate a suite of tricky Excel test files for stress-testing the parser.

Each file targets a specific real-world layout pattern that is known to trip
up naive header detection or data extraction. Run from the project root:

    python scripts/generate_test_files.py

Output: test_files/*.xlsx
"""

from pathlib import Path

import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

OUT_DIR = Path(__file__).parent.parent / "test_files"
OUT_DIR.mkdir(exist_ok=True)


# ─────────────────────────────────────────────
# 1. Multi-level merged column headers
#    Classic pivot-export style: a top header row spans groups of columns,
#    a second row has individual column names underneath.
# ─────────────────────────────────────────────
def make_multi_level_headers():
    wb = Workbook()
    ws = wb.active
    ws.title = "Revenue"

    ws["A1"] = "Regional Sales Report Q1-Q2 2024"
    ws.merge_cells("A1:I1")
    ws["A1"].font = Font(bold=True, size=14)

    # Group header row (merged)
    ws["A2"] = "Product"
    ws["B2"] = "Q1"
    ws.merge_cells("B2:D2")
    ws["E2"] = "Q2"
    ws.merge_cells("E2:G2")
    ws["H2"] = "Total"
    ws.merge_cells("H2:I2")
    for cell in ["A2", "B2", "E2", "H2"]:
        ws[cell].font = Font(bold=True)
        ws[cell].alignment = Alignment(horizontal="center")

    # Individual column headers
    headers = ["Product", "North", "South", "East", "North", "South", "East", "H1 Total", "YoY %"]
    for i, h in enumerate(headers, 1):
        ws.cell(row=3, column=i, value=h).font = Font(bold=True)

    # Data rows
    rows = [
        ("Enterprise Server Pro", 120000, 85000, 95000, 130000, 90000, 102000, None, None),
        ("CloudSync License",     45000,  32000, 38000, 52000,  36000, 41000,  None, None),
        ("Basic Support Plan",    28000,  19000, 22000, 31000,  21000, 25000,  None, None),
        ("Analytics Suite",       67000,  51000, 59000, 74000,  55000, 63000,  None, None),
        ("Dedicated Hosting",     33000,  24000, 28000, 37000,  27000, 31000,  None, None),
    ]
    for r, row in enumerate(rows, 4):
        for c, val in enumerate(row, 1):
            if val is not None:
                ws.cell(row=r, column=c, value=val)
            elif c == 8:
                # H1 Total formula
                ws.cell(row=r, column=c, value=sum(row[1:7]))
            elif c == 9:
                # YoY % as string
                pct = round((sum(row[4:7]) - sum(row[1:4])) / sum(row[1:4]) * 100, 1)
                ws.cell(row=r, column=c, value=f"{pct}%")

    # Grand total row
    ws.cell(row=9, column=1, value="GRAND TOTAL").font = Font(bold=True)
    for c in range(2, 9):
        col_letter = get_column_letter(c)
        ws.cell(row=9, column=c, value=f"=SUM({col_letter}4:{col_letter}8)")
        ws.cell(row=9, column=c).font = Font(bold=True)

    wb.save(OUT_DIR / "multi_level_headers.xlsx")
    print("✓ multi_level_headers.xlsx")


# ─────────────────────────────────────────────
# 2. Two tables stacked vertically on one sheet
#    Common in financial models: summary table at top, detail table below,
#    separated by a gap row. Parser must find the right table boundary.
# ─────────────────────────────────────────────
def make_stacked_tables():
    wb = Workbook()
    ws = wb.active
    ws.title = "Finance"

    # Table 1: Department summary
    ws["A1"] = "Department Summary"
    ws["A1"].font = Font(bold=True, size=12)

    headers1 = ["Department", "Headcount", "Q1 Budget", "Q1 Actual", "Variance"]
    for i, h in enumerate(headers1, 1):
        ws.cell(row=2, column=i, value=h).font = Font(bold=True)

    summary = [
        ("Engineering",  42, 1200000, 1187000, None),
        ("Marketing",    18,  480000,  512000, None),
        ("Sales",        31,  750000,  741000, None),
        ("Operations",   15,  320000,  308000, None),
    ]
    for r, row in enumerate(summary, 3):
        for c, val in enumerate(row, 1):
            if val is not None:
                ws.cell(row=r, column=c, value=val)
            else:
                ws.cell(row=r, column=c, value=row[3] - row[2])

    # Gap row (intentionally blank)
    # Row 7 is blank

    # Table 2: Headcount detail
    ws["A8"] = "Headcount Detail"
    ws["A8"].font = Font(bold=True, size=12)

    headers2 = ["Employee ID", "Name", "Department", "Role", "Salary", "Start Date"]
    for i, h in enumerate(headers2, 1):
        ws.cell(row=9, column=i, value=h).font = Font(bold=True)

    employees = [
        ("EMP-001", "Sarah Chen",      "Engineering", "Staff Engineer",    165000, "2021-03-15"),
        ("EMP-002", "Marcus Johnson",  "Engineering", "Senior Engineer",   145000, "2020-07-01"),
        ("EMP-003", "Emily Rodriguez", "Sales",       "Account Executive", 95000,  "2019-11-20"),
        ("EMP-004", "David Kim",       "Marketing",   "Growth Manager",    110000, "2022-01-10"),
        ("EMP-005", "Lisa Thompson",   "Sales",       "SDR",               72000,  "2023-06-01"),
        ("EMP-006", "James Park",      "Operations",  "Ops Analyst",       88000,  "2021-09-15"),
        ("EMP-007", "Nina Patel",      "Engineering", "Principal Eng",     195000, "2018-04-20"),
        ("EMP-008", "Carlos Mendez",   "Marketing",   "Brand Manager",     105000, "2022-08-01"),
    ]
    for r, row in enumerate(employees, 10):
        for c, val in enumerate(row, 1):
            ws.cell(row=r, column=c, value=val)

    wb.save(OUT_DIR / "stacked_tables.xlsx")
    print("✓ stacked_tables.xlsx")


# ─────────────────────────────────────────────
# 3. Aggressive embedded subtotals
#    Every region group has its own subtotal row inline with data.
#    Grand total at the bottom. Classic sales report export from ERP systems.
# ─────────────────────────────────────────────
def make_embedded_subtotals():
    wb = Workbook()
    ws = wb.active
    ws.title = "Sales Pipeline"

    ws["A1"] = "Global Pipeline Report"
    ws.merge_cells("A1:F1")
    ws["A1"].font = Font(bold=True, size=13)

    ws["A2"] = "As of Q2 2024 — Confidential"
    ws.merge_cells("A2:F2")

    headers = ["Opportunity", "Owner", "Stage", "Close Date", "Amount", "Probability"]
    for i, h in enumerate(headers, 1):
        ws.cell(row=3, column=i, value=h).font = Font(bold=True)

    data = [
        # Americas
        ("AMER - Acme Corp",       "Sarah Chen",      "Proposal",  "2024-07-15", 285000, "75%"),
        ("AMER - TechCorp",        "Sarah Chen",      "Negotiation","2024-06-30", 142000, "90%"),
        ("AMER - StartupXYZ",      "Marcus Johnson",  "Discovery", "2024-08-20", 65000,  "30%"),
        ("AMER - MegaCo",          "Marcus Johnson",  "Proposal",  "2024-07-01", 520000, "60%"),
        ("Americas Subtotal", "", "", "", 1012000, ""),   # subtotal row
        # EMEA
        ("EMEA - GlobalBank",      "Emily Rodriguez", "Closed Won","2024-05-01", 380000, "100%"),
        ("EMEA - EuroRetail",      "Emily Rodriguez", "Negotiation","2024-06-15", 210000, "85%"),
        ("EMEA - ManufacturingCo", "David Kim",       "Proposal",  "2024-07-30", 175000, "65%"),
        ("EMEA Subtotal",    "", "", "", 765000, ""),     # subtotal row
        # APAC
        ("APAC - AsiaTech",        "Lisa Thompson",   "Discovery", "2024-09-01", 95000,  "25%"),
        ("APAC - PacificGroup",    "Lisa Thompson",   "Proposal",  "2024-07-15", 445000, "70%"),
        ("APAC - KoreaFinance",    "James Park",      "Negotiation","2024-06-20", 320000, "80%"),
        ("APAC Subtotal",    "", "", "", 860000, ""),     # subtotal row
        # Grand total
        ("GRAND TOTAL",      "", "", "", 2637000, ""),
    ]

    bold_rows = {"Americas Subtotal", "EMEA Subtotal", "APAC Subtotal", "GRAND TOTAL"}
    for r, row in enumerate(data, 4):
        for c, val in enumerate(row, 1):
            cell = ws.cell(row=r, column=c, value=val)
            if row[0] in bold_rows:
                cell.font = Font(bold=True)

    wb.save(OUT_DIR / "embedded_subtotals.xlsx")
    print("✓ embedded_subtotals.xlsx")


# ─────────────────────────────────────────────
# 4. Mixed currency + percentage columns
#    Multi-currency expense report with USD, EUR, GBP values in the same
#    column. Percentage approval rates. Tests hint detection and aggregation
#    guards.
# ─────────────────────────────────────────────
def make_mixed_currency():
    wb = Workbook()
    ws = wb.active
    ws.title = "Expenses"

    ws["A1"] = "Global Expense Report — H1 2024"
    ws.merge_cells("A1:G1")
    ws["A1"].font = Font(bold=True, size=12)

    ws["A2"] = "(All amounts in local currency — do not aggregate across currencies)"
    ws.merge_cells("A2:G2")

    headers = ["Expense ID", "Employee", "Category", "Currency", "Amount", "Approved", "Approval Rate"]
    for i, h in enumerate(headers, 1):
        ws.cell(row=3, column=i, value=h).font = Font(bold=True)

    expenses = [
        ("EXP-001", "Sarah Chen",      "Travel",        "USD", "$4,250",   "Yes", "98%"),
        ("EXP-002", "Marcus Johnson",  "Software",      "USD", "$1,800",   "Yes", "100%"),
        ("EXP-003", "Emily Rodriguez", "Travel",        "EUR", "€3,100",   "Yes", "95%"),
        ("EXP-004", "David Kim",       "Entertainment", "GBP", "£850",     "No",  "72%"),
        ("EXP-005", "Lisa Thompson",   "Equipment",     "USD", "$6,400",   "Yes", "100%"),
        ("EXP-006", "James Park",      "Travel",        "EUR", "€2,750",   "Yes", "91%"),
        ("EXP-007", "Nina Patel",      "Software",      "GBP", "£1,200",   "Yes", "100%"),
        ("EXP-008", "Carlos Mendez",   "Travel",        "USD", "$3,900",   "Pending", "85%"),
        ("EXP-009", "Sarah Chen",      "Equipment",     "EUR", "€5,600",   "Yes", "97%"),
        ("EXP-010", "Emily Rodriguez", "Entertainment", "GBP", "£420",     "No",  "68%"),
        ("EXP-011", "Marcus Johnson",  "Travel",        "USD", "$2,100",   "Yes", "100%"),
        ("EXP-012", "David Kim",       "Software",      "EUR", "€980",     "Yes", "99%"),
    ]
    for r, row in enumerate(expenses, 4):
        for c, val in enumerate(row, 1):
            ws.cell(row=r, column=c, value=val)

    # Summary footer
    ws["A17"] = "Note: FX rates as of 2024-06-30. USD/EUR: 1.08, USD/GBP: 1.27"
    ws.merge_cells("A17:G17")

    wb.save(OUT_DIR / "mixed_currency.xlsx")
    print("✓ mixed_currency.xlsx")


# ─────────────────────────────────────────────
# 5. KPI dashboard — wide, sparse, metric rows
#    One row per KPI metric, multiple period columns. No traditional
#    "column of IDs" — the metric name IS the identifier.
# ─────────────────────────────────────────────
def make_kpi_dashboard():
    wb = Workbook()
    ws = wb.active
    ws.title = "KPIs"

    ws["A1"] = "Company KPI Dashboard"
    ws.merge_cells("A1:H1")
    ws["A1"].font = Font(bold=True, size=14)

    ws["A2"] = "Updated: June 30, 2024"
    ws["A3"] = "Source: Internal BI — not for external distribution"

    headers = ["Metric", "Unit", "Q1 2023", "Q2 2023", "Q3 2023", "Q4 2023", "Q1 2024", "Q2 2024"]
    for i, h in enumerate(headers, 1):
        ws.cell(row=4, column=i, value=h).font = Font(bold=True)

    kpis = [
        ("ARR",               "$M",  8.2,   9.1,   10.4,  11.8,  13.2,  14.9),
        ("MRR",               "$M",  0.683, 0.758, 0.867, 0.983, 1.100, 1.242),
        ("New Logos",         "#",   12,    15,    18,    22,    19,    24),
        ("Churn Rate",        "%",   "2.1%","1.8%","2.4%","1.9%","1.6%","1.4%"),
        ("NRR",               "%",   "108%","111%","114%","116%","118%","121%"),
        ("CAC",               "$",   8400,  7900,  7200,  6800,  7100,  6500),
        ("LTV",               "$",   42000, 45000, 48000, 52000, 55000, 59000),
        ("LTV:CAC",           "x",   5.0,   5.7,   6.7,   7.6,   7.7,   9.1),
        ("Payback Period",    "mo",  18,    16,    15,    14,    14,    13),
        ("Gross Margin",      "%",   "71%", "72%", "73%", "74%", "74%", "75%"),
        ("Win Rate",          "%",   "28%", "31%", "33%", "35%", "32%", "38%"),
        ("Avg Deal Size",     "$",   23750, 24100, 25800, 26900, 27400, 29200),
        ("Sales Cycle",       "days",47,    44,    42,    41,    43,    39),
        ("Support Tickets",   "#",   234,   198,   215,   187,   201,   176),
        ("CSAT Score",        "/10", 8.2,   8.4,   8.3,   8.6,   8.7,   8.9),
    ]

    for r, row in enumerate(kpis, 5):
        for c, val in enumerate(row, 1):
            ws.cell(row=r, column=c, value=val)

    wb.save(OUT_DIR / "kpi_dashboard.xlsx")
    print("✓ kpi_dashboard.xlsx")


# ─────────────────────────────────────────────
# 6. Blank header columns + formula-derived totals
#    Sheet with intentionally unnamed columns (like the col_7 case) plus
#    a column whose values are derived from formulas — tests both blank
#    header inference and formula value reading.
# ─────────────────────────────────────────────
def make_blank_headers_and_formulas():
    wb = Workbook()
    ws = wb.active
    ws.title = "Orders"

    ws["A1"] = "Order Management System Export"
    ws.merge_cells("A1:H1")
    ws["A1"].font = Font(bold=True)

    ws["A2"] = "Exported: 2024-06-30 09:15 UTC"

    headers = ["Order ID", "Customer", "Product", "Qty", "Unit Price", "Discount %", "Tax Rate", None, None]
    for i, h in enumerate(headers, 1):
        if h is not None:
            ws.cell(row=3, column=i, value=h).font = Font(bold=True)
        # columns H (8) and I (9) intentionally left blank — formula-derived

    orders = [
        ("ORD-1001", "Acme Corp",     "Server Pro",   5,  2999.99, "10%", "8.5%"),
        ("ORD-1002", "TechStart",     "CloudSync",    12, 149.99,  "0%",  "8.5%"),
        ("ORD-1003", "GlobalBank",    "Analytics",    3,  3999.99, "15%", "0%"),
        ("ORD-1004", "MegaRetail",    "Support Plan", 50, 299.99,  "20%", "8.5%"),
        ("ORD-1005", "EuroFinance",   "Hosting",      8,  599.99,  "5%",  "0%"),
        ("ORD-1006", "AsiaGroup",     "Server Pro",   2,  2999.99, "0%",  "0%"),
        ("ORD-1007", "LocalBiz",      "CloudSync",    25, 149.99,  "10%", "8.5%"),
        ("ORD-1008", "NatGovAgency",  "Analytics",    10, 3999.99, "25%", "0%"),
    ]

    for r, row in enumerate(orders, 4):
        for c, val in enumerate(row, 1):
            ws.cell(row=r, column=c, value=val)
        # Col H: Subtotal (Qty * Unit Price * (1 - Discount))
        ws.cell(row=r, column=8, value=f"=D{r}*E{r}*(1-SUBSTITUTE(F{r},\"%\",\"\")/100)")
        # Col I: Total with tax
        ws.cell(row=r, column=9, value=f"=H{r}*(1+SUBSTITUTE(G{r},\"%\",\"\")/100)")

    # Grand total row
    ws["A12"] = "TOTAL"
    ws["A12"].font = Font(bold=True)
    ws["H12"] = "=SUM(H4:H11)"
    ws["I12"] = "=SUM(I4:I11)"
    ws["H12"].font = Font(bold=True)
    ws["I12"].font = Font(bold=True)

    wb.save(OUT_DIR / "blank_headers_formulas.xlsx")
    print("✓ blank_headers_formulas.xlsx")


# ─────────────────────────────────────────────
# 7. Multi-sheet workbook with real joins
#    Like company_data.xlsx but with nastier cross-sheet linking:
#    different column naming conventions, case mismatches, one-to-many.
# ─────────────────────────────────────────────
def make_messy_joins():
    wb = Workbook()

    # Sheet 1: Invoices
    ws1 = wb.active
    ws1.title = "Invoices"
    ws1.append(["Invoice #", "CLIENT_ID", "Issue Date", "Due Date", "Line Items", "Subtotal", "Tax", "TOTAL"])
    invoices = [
        ("INV-2024-001", "C001", "2024-01-15", "2024-02-15", 3, 45000, 3825,  48825),
        ("INV-2024-002", "C003", "2024-01-22", "2024-02-22", 1, 12000, 0,     12000),
        ("INV-2024-003", "C002", "2024-02-01", "2024-03-01", 5, 87500, 7437,  94937),
        ("INV-2024-004", "C001", "2024-02-14", "2024-03-14", 2, 31000, 2635,  33635),
        ("INV-2024-005", "C004", "2024-03-01", "2024-04-01", 4, 62000, 5270,  67270),
        ("INV-2024-006", "C002", "2024-03-15", "2024-04-15", 1, 18500, 1572,  20072),
        ("INV-2024-007", "C005", "2024-04-01", "2024-05-01", 6, 124000,0,     124000),
        ("INV-2024-008", "C003", "2024-04-15", "2024-05-15", 2, 29000, 2465,  31465),
    ]
    for row in invoices:
        ws1.append(row)

    # Sheet 2: Clients (note: uses "client_id" not "CLIENT_ID" — case mismatch)
    ws2 = wb.create_sheet("Clients")
    ws2.append(["client_id", "Company Name", "Industry", "Account Manager", "Contract Tier", "Annual Value"])
    clients = [
        ("C001", "Acme Corporation",  "Manufacturing", "Sarah Chen",      "Enterprise", 480000),
        ("C002", "TechStart Inc",     "SaaS",          "Marcus Johnson",  "Growth",     210000),
        ("C003", "GlobalBank Ltd",    "Finance",       "Emily Rodriguez", "Enterprise", 380000),
        ("C004", "RetailGiant Co",    "Retail",        "David Kim",       "Standard",   95000),
        ("C005", "GovAgency Federal", "Government",    "Lisa Thompson",   "Enterprise", 750000),
    ]
    for row in clients:
        ws2.append(row)

    # Sheet 3: Payments (links to invoices by invoice number — semantic match)
    ws3 = wb.create_sheet("Payments")
    ws3.append(["Payment ID", "Related Invoice", "Payment Date", "Method", "Amount Received", "Outstanding"])
    payments = [
        ("PAY-001", "INV-2024-001", "2024-02-10", "Wire",   48825, 0),
        ("PAY-002", "INV-2024-002", "2024-02-20", "ACH",    12000, 0),
        ("PAY-003", "INV-2024-003", "2024-03-05", "Wire",   50000, 44937),
        ("PAY-004", "INV-2024-003", "2024-03-28", "Wire",   44937, 0),
        ("PAY-005", "INV-2024-004", "2024-03-12", "Check",  33635, 0),
        ("PAY-006", "INV-2024-005", "2024-04-15", "Wire",   67270, 0),
        ("PAY-007", "INV-2024-007", "2024-05-01", "Wire",   124000,0),
    ]
    for row in payments:
        ws3.append(row)

    wb.save(OUT_DIR / "messy_joins.xlsx")
    print("✓ messy_joins.xlsx")


if __name__ == "__main__":
    print(f"Generating test files in {OUT_DIR}/\n")
    make_multi_level_headers()
    make_stacked_tables()
    make_embedded_subtotals()
    make_mixed_currency()
    make_kpi_dashboard()
    make_blank_headers_and_formulas()
    make_messy_joins()
    print(f"\nDone. {len(list(OUT_DIR.glob('*.xlsx')))} files in {OUT_DIR}/")
