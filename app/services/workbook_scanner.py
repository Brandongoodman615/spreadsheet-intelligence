import logging
from pathlib import Path

import openpyxl
import pandas as pd
from pydantic import BaseModel

logger = logging.getLogger(__name__)


class MergedRegion(BaseModel):
    range_str: str
    top_left_row: int       # 0-indexed
    top_left_col: int       # 0-indexed
    top_left_value: str | None


class SheetScan(BaseModel):
    bold_rows: list[int]            # 0-indexed rows where majority of non-empty cells are bold
    merged_regions: list[MergedRegion]
    named_ranges: list[str]         # range names that reference this sheet
    print_area: str | None


class WorkbookScan(BaseModel):
    sheets: dict[str, SheetScan]


def scan_workbook(wb: openpyxl.Workbook) -> WorkbookScan:
    """
    Extract rich structural metadata from an already-open openpyxl Workbook.

    Captures formatting signals (bold rows, merged regions, named ranges, print
    areas) that are discarded by pandas. This metadata is passed to the structure
    analyzer so the LLM has explicit layout signals rather than guessing from a
    flat cell dump.

    Accepts an open Workbook so the file is not read a second time.
    Returns an empty WorkbookScan on failure so the rest of the pipeline
    degrades gracefully to the existing heuristic behavior.
    """
    try:
        sheet_scans: dict[str, SheetScan] = {}

        # Build a mapping of sheet title → range names for named range lookup
        named_ranges_by_sheet: dict[str, list[str]] = {ws.title: [] for ws in wb.worksheets}
        for defined_name in wb.defined_names:
            try:
                for sheet_title, _ in defined_name.destinations:
                    if sheet_title in named_ranges_by_sheet:
                        named_ranges_by_sheet[sheet_title].append(defined_name.name)
            except Exception:
                continue

        for ws in wb.worksheets:
            sheet_scans[ws.title] = _scan_sheet(ws, named_ranges_by_sheet.get(ws.title, []))

        return WorkbookScan(sheets=sheet_scans)

    except Exception as exc:
        logger.warning("Workbook pre-scan failed (%s) — structure analyzer will run without metadata", exc)
        return WorkbookScan(sheets={})


def _scan_sheet(ws, named_ranges: list[str]) -> SheetScan:
    bold_rows = _extract_bold_rows(ws)
    merged_regions = _extract_merged_regions(ws)
    print_area = _extract_print_area(ws)
    return SheetScan(
        bold_rows=bold_rows,
        merged_regions=merged_regions,
        named_ranges=named_ranges,
        print_area=print_area,
    )


def _extract_bold_rows(ws) -> list[int]:
    """
    Return 0-indexed row numbers where the majority of non-empty cells are bold.

    Scans the entire sheet so subtotal rows and section headers embedded deep
    in the data are captured, not just header candidates near the top.
    """
    bold_rows = []
    for row in ws.iter_rows():
        non_empty = [cell for cell in row if cell.value is not None]
        if not non_empty:
            continue
        bold_count = sum(
            1 for cell in non_empty
            if cell.font and cell.font.bold
        )
        if bold_count / len(non_empty) >= 0.5:
            bold_rows.append(row[0].row - 1)  # convert to 0-indexed
    return bold_rows


def _extract_merged_regions(ws) -> list[MergedRegion]:
    """
    Return all merged cell regions with their anchor cell value.

    Merged cells spanning multiple columns at the top of a sheet are almost
    always title rows. Merged cells spanning multiple rows within a column
    are group headers (e.g. "Americas" covering 3 data rows).
    """
    regions = []
    for merged_range in ws.merged_cells.ranges:
        top_left_row = merged_range.min_row - 1   # 0-indexed
        top_left_col = merged_range.min_col - 1   # 0-indexed
        anchor = ws.cell(merged_range.min_row, merged_range.min_col)
        value = str(anchor.value).strip() if anchor.value is not None else None
        regions.append(MergedRegion(
            range_str=str(merged_range),
            top_left_row=top_left_row,
            top_left_col=top_left_col,
            top_left_value=value,
        ))
    return regions


def _extract_print_area(ws) -> str | None:
    """Return the print area with $ signs stripped, or None if not set."""
    if not ws.print_area:
        return None
    return ws.print_area.replace("$", "")


def build_raw_dataframe(ws) -> pd.DataFrame:
    """
    Build a raw DataFrame from an openpyxl worksheet using iter_rows.

    This replaces pd.read_excel for the initial load so the file is read only
    once (by scan_workbook) rather than twice. Values are extracted as-is;
    all cleaning happens downstream in _extract_table.
    """
    data = [[cell.value for cell in row] for row in ws.iter_rows()]
    return pd.DataFrame(data)
