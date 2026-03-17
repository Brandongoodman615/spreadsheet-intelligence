import re
from pathlib import Path
import pandas as pd
import openpyxl
from app.services.sheet_structure_analyzer import analyze_sheet_structure


def load_workbook(path: Path) -> dict[str, pd.DataFrame]:
    """
    Load all sheets from an .xlsx file into a dict of DataFrames.

    Uses data_only=True so formula cells return their last cached value.
    Does NOT recalculate formulas — see README known limitations.
    """
    frames: dict[str, pd.DataFrame] = {}

    raw = pd.read_excel(path, sheet_name=None, engine="openpyxl", header=None)

    for sheet_name, df in raw.items():
        cleaned = _extract_table(df, sheet_name)
        if cleaned is not None:
            frames[sheet_name] = cleaned

    return frames


def _extract_table(df: pd.DataFrame, sheet_name: str) -> pd.DataFrame | None:
    """
    Find the best header row and return clean tabular data.

    Uses the LLM to detect sheet structure (header row, data start, rows to skip)
    with a heuristic fallback. Handles any layout: leading title rows, merged cell
    group headers, embedded subtotals, trailing metadata, duplicate column names,
    currency symbols, and percentage strings.
    """
    if df.empty:
        return None

    # Drop fully empty rows and columns first so the LLM sees a clean sample.
    df = df.dropna(how="all").dropna(axis=1, how="all").reset_index(drop=True)

    if df.empty:
        return None

    # Ask the LLM to identify the header row, data start row, and any rows
    # to skip (subtotals, grand totals, metadata rows embedded in the data).
    structure = analyze_sheet_structure(df, sheet_name)

    header_idx = structure.header_row
    raw_cols = [
        str(df.loc[header_idx, c]).strip() if pd.notna(df.loc[header_idx, c]) else f"col_{i}"
        for i, c in enumerate(df.columns)
    ]

    # Deduplicate column names to avoid ambiguous DataFrame[col] lookups
    seen: dict[str, int] = {}
    deduped = []
    for name in raw_cols:
        if name in seen:
            seen[name] += 1
            deduped.append(f"{name}_{seen[name]}")
        else:
            seen[name] = 0
            deduped.append(name)

    df.columns = deduped

    # Keep only rows from data_start_row onward, excluding skip_rows
    data_rows = df.loc[structure.data_start_row:]
    if structure.skip_rows:
        data_rows = data_rows.drop(
            index=[r for r in structure.skip_rows if r in data_rows.index],
            errors="ignore",
        )
    df = data_rows.reset_index(drop=True)

    # Drop rows that are entirely empty
    df = df.dropna(how="all")

    if df.empty:
        return None

    # Drop units/descriptor rows that appear at the top of the data block.
    # These are rows immediately below the header where all non-null values are
    # parenthetical labels like "(USD)", "(Local)", "(Expected)" — a common Excel
    # convention for annotating column units or sub-headers.
    df = _strip_leading_descriptor_rows(df)

    if df.empty:
        return None

    # Drop any trailing metadata rows the LLM may have missed.
    # This is a lightweight safety net for footers that appear after the last
    # data row and were outside the LLM's sample window.
    df = _strip_trailing_metadata(df)

    if df.empty:
        return None

    # Forward-fill sparse columns.
    # Some sheets use merged cells for group headers (e.g. "Americas" spanning 3 rows,
    # "SALES" spanning a budget section). After parsing, these appear as NaN in all
    # but the first row of the group. Forward-fill restores the grouping so SQL
    # GROUP BY and WHERE work correctly.
    df = _forward_fill_sparse_columns(df)

    # Clean string values: strip currency symbols, annotation characters, and
    # comma formatting so numeric columns can be cast correctly in SQL.
    for col in df.select_dtypes(include="object").columns:
        df[col] = df[col].map(_clean_cell_value)

    # Normalize percent columns to uniform decimal floats BEFORE type coercion.
    # Handles mixed columns where some cells are "75%" strings and others are
    # already-decimal floats (e.g. 0.65). Normalizes everything to [0, 1] range.
    df = _normalize_percent_columns(df)

    # Coerce types: try numeric conversion on object columns.
    # Only apply if conversion introduces no new NaN values (mirrors old errors="ignore").
    for col in df.select_dtypes(include="object").columns:
        converted = pd.to_numeric(df[col], errors="coerce")
        if converted.isna().sum() == df[col].isna().sum():
            df[col] = converted

    return df


def _strip_leading_descriptor_rows(df: pd.DataFrame) -> pd.DataFrame:
    """
    Remove units/descriptor rows at the top of the data block.

    Targets rows immediately below the header where every non-null value is a
    parenthetical label — e.g. "(USD)", "(Local)", "(Expected)", "(Revised)".
    This is a standard Excel convention for annotating column units or sub-headers
    and is not data.

    Only strips contiguous matching rows from the top; stops at the first row
    that contains any non-parenthetical value.
    """
    paren_re = re.compile(r'^\(.*\)$')

    drop_up_to = 0
    for i, (_, row) in enumerate(df.iterrows()):
        non_null = row.dropna()
        if non_null.empty:
            drop_up_to = i + 1
            continue
        all_parens = all(paren_re.match(str(v).strip()) for v in non_null)
        if all_parens:
            drop_up_to = i + 1
        else:
            break

    return df.iloc[drop_up_to:].reset_index(drop=True)


def _strip_trailing_metadata(df: pd.DataFrame) -> pd.DataFrame:
    """
    Remove footer rows that appear after the main data block.

    Detects rows where >60% of columns are null AND the first non-null value
    is a string that looks like a label or annotation rather than data.
    Strips from the bottom up until a clean data row is found.
    """
    if df.empty:
        return df

    col_count = len(df.columns)
    last_good = len(df)

    for i in range(len(df) - 1, -1, -1):
        row = df.iloc[i]
        null_ratio = row.isna().sum() / col_count
        if null_ratio < 0.6:
            break
        # High null ratio — check if first non-null looks like metadata
        non_null = row.dropna()
        if non_null.empty:
            last_good = i
            continue
        first_val = str(non_null.iloc[0]).strip()
        # Metadata patterns: starts with common label prefixes, or is very long prose
        if re.match(r"^(data as of|source:|note:|w\d+\s*=|updated|last|version|\*)", first_val, re.IGNORECASE) \
                or len(first_val) > 60:
            last_good = i
        else:
            break

    return df.iloc[:last_good].reset_index(drop=True)


def _normalize_percent_columns(df: pd.DataFrame) -> pd.DataFrame:
    """
    Convert percent-string columns to uniform decimal floats.

    Handles mixed columns where some cells are "75%" strings and others are
    already-decimal floats (e.g. 0.65 for 65%). Normalizes everything to [0, 1]
    so queries can use the column directly without SQL REPLACE/divide gymnastics.

    A column qualifies if ≥50% of its non-null string values end with '%'.
    """
    for col in df.select_dtypes(include="object").columns:
        non_null = df[col].dropna()
        if non_null.empty:
            continue

        str_vals = non_null[non_null.apply(lambda x: isinstance(x, str))]
        if str_vals.empty:
            continue

        pct_frac = str_vals.astype(str).str.match(r'^-?\d+(\.\d+)?%$').sum() / len(non_null)
        if pct_frac < 0.5:
            continue

        def _to_decimal(v, _col=col):
            if pd.isna(v):
                return v
            if isinstance(v, (int, float)):
                f = float(v)
                # Float already in [0, 1] is a decimal; larger values are raw percentages
                return f if abs(f) <= 1.0 else f / 100.0
            s = str(v).strip()
            if s.endswith('%'):
                try:
                    return float(s[:-1]) / 100.0
                except ValueError:
                    return None
            try:
                f = float(s)
                return f if abs(f) <= 1.0 else f / 100.0
            except ValueError:
                return v

        df[col] = df[col].map(_to_decimal)

    return df


def _forward_fill_sparse_columns(df: pd.DataFrame) -> pd.DataFrame:
    """
    Forward-fill columns that are sparsely populated due to merged cell groups.

    Only fills columns where:
    - dtype is object (text grouping columns, not numeric data)
    - At least 30% of values are null (genuinely sparse, not just missing data)
    - The non-null values repeat in a pattern consistent with group headers

    This handles sheets like Revenue (Region: Americas/EMEA/APAC spanning rows)
    and Budgets (Department: SALES/MARKETING/ENGINEERING spanning line items).
    """
    for col in df.select_dtypes(include="object").columns:
        null_ratio = df[col].isna().sum() / len(df)
        if null_ratio >= 0.3:
            df[col] = df[col].ffill().infer_objects(copy=False)
    return df


def _clean_cell_value(val):
    """
    Clean a single cell value for reliable numeric coercion.

    Handles:
    - Currency symbols: $1,234.56 → 1234.56, €88,500 → 88500
    - Percentage strings: "75%" → "75" (dtype hint preserved in schema)
    - Comma-formatted numbers: "1,234,567" → "1234567"
    - Annotated values: "158,300*" → "158300"
    - Sub-header labels like "(USD)", "(Local)" → kept as-is (non-numeric)
    - Whitespace-padded hierarchy strings: "  Finance" → "Finance"
    """
    if not isinstance(val, str):
        return val

    stripped = val.strip()

    # Strip leading/trailing whitespace from hierarchy-indented labels
    if stripped != val:
        val = stripped

    # Normalize ALL-CAPS label words to Title Case (e.g. ENGINEERING → Engineering).
    # Common in Excel for department/category headers. Only applies to purely alphabetic
    # words — leaves codes (EMP-001, USD, CAT-HW) and mixed-case strings unchanged.
    if stripped.isupper() and stripped.replace(' ', '').isalpha():
        return stripped.title()

    # Remove annotation suffix (* means estimated/restated)
    cleaned = re.sub(r'\*+$', '', stripped)

    # Strip currency symbols and formatting
    cleaned = re.sub(r'^[€£¥₹₩]', '', cleaned)
    cleaned = re.sub(r'^\$', '', cleaned)

    # Remove comma formatting in numbers (e.g. "1,234,567")
    if re.match(r'^-?[\d,]+(\.\d+)?$', cleaned):
        cleaned = cleaned.replace(',', '')

    return cleaned if cleaned != '' else val


def has_formula_cells(path: Path) -> bool:
    """Check if any cell in the workbook contains a formula."""
    wb = openpyxl.load_workbook(path, data_only=False)
    for sheet in wb.worksheets:
        for row in sheet.iter_rows():
            for cell in row:
                if isinstance(cell.value, str) and cell.value.startswith("="):
                    return True
    return False
